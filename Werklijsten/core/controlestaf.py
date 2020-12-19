import openpyxl as xl
from openpyxl import Workbook
import re
import itertools
import datetime

# TODO: alles omwerken naar dictionary !
# TODO: Leesbaarheid moet beter

MAX_LIST = 0
MIN_STAFF = 0
MAX_STAFF = 0
MAX_VIRGA = 0
MIN_SAL = 0
MAX_SAL = 0


def main(wb, weekkeuze):
    control_data = list()
    weken = kiesweken(wb, weekkeuze)
    li = maaklijsten(wb)
    dagschemas = maakalledagschemas(weken, li)
    aanwezigen_salvator = check_aanwezigen_sal_opdatum(weken)
    aanwezigen_cardio_virga = check_aanwezigen_virga_var_opdatum(weken, 7)

    # voer de verschillende controles uit
    control_data += controleeraflossing(dagschemas, li)
    control_data += controleerchefok(dagschemas)
    control_data += checkdaywithoutfunction(dagschemas)
    control_data += checkwachten(weken, li)
    control_data += checkrecup(dagschemas)
    control_data += check_v3(dagschemas)
    control_data += check_locoreg_anesthesist_sal(aanwezigen_salvator, li)
    control_data += check_cardioanesthesist_vj(aanwezigen_cardio_virga, li)

    return control_data


def kiesweken(workbook, weekkeuze):
    sheets = [workbook['WEEK1'], workbook['WEEK2'], workbook['WEEK3'], workbook['WEEK4']]

    if weekkeuze == 'all':
        sheets_to_use = sheets
    else:
        sheets_to_use = [workbook[weekkeuze]]
    return sheets_to_use


def maaklijsten(workbook):
    # bijkomende functie maken !
    # maakt de lijsten aan om in andere functies te gebruiken, output is een lijst dictionaries
    # lijsten[0] = staff, lijsten[1]= intensieve, lijsten[2]= posities virga, lijsten[3] = posities salvator
    lijst = {"staff": [], "intensievist": [], "posities_virga": [], "posities_salvator": [], "cardio_anesthesist": [],
             "assistent_anesthesie": [], "assistent_ite": [], "locoreg_anesthesist": []}

    # Maak lijst Intensivisten
    sheet = workbook['ARTSEN']
    for i in range(2, len(sheet['B']) + 1):
        if sheet.cell(i, 2).value is not None:
            lijst["intensievist"].append(sheet.cell(i, 2).value)

    # Maak lijst stafleden
    for i in range(2, len(sheet['A']) + 1):
        if sheet.cell(i, 1).value is not None:
            lijst["staff"].append(sheet.cell(i, 1).value)

    # Maak lijst cardio_anesthesist
    for i in range(2, len(sheet['C']) + 1):
        if sheet.cell(i, 3).value is not None:
            lijst["cardio_anesthesist"].append(sheet.cell(i, 3).value)

    # Maak lijst assistenten_anesthesie
    for i in range(2, len(sheet['G']) + 1):
        if sheet.cell(i, 7).value is not None:
            lijst["assistent_anesthesie"].append(sheet.cell(i, 7).value)

    # Maak lijst assistent_ite
    for i in range(2, len(sheet['H']) + 1):
        if sheet.cell(i, 8).value is not None:
            lijst["assistent_anesthesie"].append(sheet.cell(i, 8).value)

    # Maak lijst locoreg_anesthesist
    for i in range(2, len(sheet['D']) + 1):
        if sheet.cell(i, 4).value is not None:
            lijst["locoreg_anesthesist"].append(sheet.cell(i, 4).value)

    # Maak een lijst van Virga Jesse posities
    sheet = workbook['WEEK4']
    setdimensions(sheet)
    for i in range(5, MAX_VIRGA + 1):
        if sheet.cell(i, 2).value is not None:
            lijst["posities_virga"].append(sheet.cell(i, 2).value)

    # Maak lijst van Salvator posities
    for i in range(MIN_SAL, MAX_SAL + 1):
        if sheet.cell(i, 2).value is not None:
            lijst["posities_salvator"].append(sheet.cell(i, 2).value)

    return lijst


def maakalledagschemas(sheets_to_use, lijsten):
    # staff is eerste lijst van lijsten, later kunnen andere lijsten toegevoegd worden
    # return is een lijst van alle dagschema's van alle stafleden over gekozen periode
    staff = lijsten["staff"]
    cumulschemas = []

    for sheet in sheets_to_use:
        setdimensions(sheet)
        for staflid in staff:
            cumulschemas.append(maakindividueledagschemas(staflid, sheet))
    flattedlijst = list(itertools.chain.from_iterable(cumulschemas))

    return flattedlijst


def maakindividueledagschemas(staflid, sheet):
    """Maak een schema per staflid per dag en return een lijst met al zijn/haar dagschemas
    Een schema bestaat steeds minstens uit: staflid, dag, datum"""

    individueelschema = []
    # doorloop alle dagen van een week
    for col in range(3, 10):

        # datum in rij 4
        # date = sheet.cell(4, col).value.date()
        temp_date = sheet.cell(4, col).value.date()
        date = temp_date.strftime("%d/%m/%Y")

        # initieer lijst met naam staflid, dag en datum
        schema = [staflid, sheet.cell(3, col).value, date]

        # loop over ale verschillende posities en diensten tot aan MAX_LIST
        for row in range(3, MAX_LIST + 1):  # 127

            # als staflid gevonden wordt dan staat zijn dienst in kolom 2
            if staflid in str(sheet.cell(row, col).value):  # conversie naar string -> iterable
                schema.append(sheet.cell(row, 2).value)

        # voeg dagschema toe aan individueel schema staflid
        individueelschema.append(schema)
    return individueelschema


def setdimensions(sheet):
    """Bepaal de dimensies van het werkblad met de cruciale posities"""
    # controleer sheet, zoek tot waar werklijst loopt en van waar tot waar de stafleden staan(BRN-WKW)
    # MAX_LIST, MIN_STAFF en MAX_STAFF worden hier bepaald
    # zoek wat de min en max rijnummers voor posities in virga en salvator zijn
    global MAX_LIST
    global MIN_STAFF
    global MAX_STAFF
    global MAX_VIRGA
    global MIN_SAL
    global MAX_SAL
    for row in range(1, sheet.max_row):
        if sheet.cell(row, 2).value == "RECUP":
            MAX_LIST = row
        if sheet.cell(row, 2).value == "Radiologie 2":
            MAX_VIRGA = row
        if sheet.cell(row, 2).value == "S-ITE 1 Corona-Coordinator Sal":
            MIN_SAL = row
        if sheet.cell(row, 2).value == "RAADPLEGING Salvator":
            MAX_SAL = row
        if sheet.cell(row, 3).value == "BRN" and row > MAX_LIST + 1:
            MIN_STAFF = row
        if sheet.cell(row, 3).value == "WKW" and row > MAX_LIST:
            MAX_STAFF = row



def checkweekend(day):
    weekend = False
    if day == "Zaterdag" or day == "Zondag":
        weekend = True
    return weekend


def checkdaywithoutfunction(schemas):
    """ weekdag zonder toewijzingen
    schema bestaat altijd uit minstens [staflid, dag, datum ], de volgende elementen zijn toewijzigingen"""
    # TODO: lijst met feestdagen importeren -> moet bestaan in python!
    result = list()

    for schema in schemas:
        if len(schema) == 3:
            if not checkweekend(schema[1]):
                result.append(f'Geen toewijzingen voor  {schema[0]} op {schema[1]} {schema[2]}')
    return result


def lookformatches(schema, matchcount, lijst, functie):
    result = ""
    matches = [x for x in schema[3:] if str(x) in lijst]
    if len(matches) > matchcount:
        zalen = " ,".join(matches)
        result += f'{functie} niet compatibel met zaal: {schema[0]} op {schema[1]} {schema[2]} ' + zalen

    return result


def controleeraflossing(schemas, lijsten):
    """komt staflid meerdere keren op zelfde afloslijst voor, of op twee afloslijsten"""
    # TODO: functie opsplitsen in meerdere functies
    result = list()
    for schema in schemas:
        if not checkweekend(schema[1]):
            # regex: start met V of S gevolgd door één of meer getallen of een woord (bv V1 inslapend)
            # meerdere keren op afloslijst
            pattern = re.compile(r'^[VS]\d+\w*')
            matches = [x for x in schema[3:] if pattern.match(str(x)) or x == 'Eerst Vertrekkende']
            if len(matches) > 1:
                posities = " ,".join(matches)
                result.append(f'Meer dan één keer op afloslijsten: {schema[0]} op {schema[2]}: ' + posities)

            # staflid op afloslijst Salvator maar ingepland Virga Jesse of alleen op afloslijst Salvator
            pattern = re.compile(r'^[S]\d\w*')
            matches = [x for x in schema[3:] if
                       pattern.match(str(x)) and (schema[3] in lijsten["posities_virga"] or len(schema) == 4)]
            if len(matches) >= 1:
                posities = " ,".join(matches)
                result.append(f'Controleer aflossing/zaaltoewijzing voor: {schema[0]} op {schema[2]}: ' + posities)

            # staflid op afloslijst Virga Jesse maar ingepland op Salvator of alleen op afloslijst Virga Jesse
            pattern = re.compile(r'^[V]\d\w*')
            matches = [x for x in schema[3:] if (pattern.match(str(x)) or
                                                 x == 'Eerst Vertrekkende') and
                       (schema[3] in lijsten["posities_salvator"] or len(schema) == 4)]
            if len(matches) >= 1:
                posities = " ,".join(matches)
                result.append(f'Controleer aflossing/zaaltoewijzing voor: {schema[0]} op {schema[2]}: ' + posities)
            if len(result) == 0:
                result.append('Geen afwijking gevonden in combinatie zaal/aflos')
    return result


def controleerchefok(schemas):
    """chef ok kan niet in bepaalde zalen staan(notcompatible)"""
    result = list()

    for schema in schemas:

        # controleer positie chef OK virga jesse
        if "Chef OK VIRGA" in schema:
            notcompatible = ["ZAAL 11", "ZAAL 12", "ITE 1", "ITE 2", "ITE 3", "ZAAL 7 CARDIO", "HYBRIDE", "EFO 2",
                             "Radiologie 1"]
            match = lookformatches(schema, 0, notcompatible, "Chef OK VIRGA")
            if match != "":
                result.append(match)

        # controleer positie chef OK Salvator
        if "Chef OK Salvator" in schema:
            notcompatible = ["OFTALMO ZAAL 11", "OFTALMO ZAAL 12", "S-ZAAL 7", "S-ZAAL 8", "S-ZAAL 9"]
            match = lookformatches(schema, 0, notcompatible, "Chef OK Salvator")
            if match != "":
                result.append(match)
        if len(result) == 0:
            result.append("Corona-Coördinator OK VJ en Chef OK Salvator in correcte zalen")
    return result


def checkwachten(weken, lijsten):
    # controleer of in de wachten minstens één intensivist aanwezig is
    result = list()
    intensieve = lijsten["intensievist"]
    for sheet in weken:
        for col in range(3, 10):
            date = sheet.cell(4, col).value.date()
            wachten = {}
            for row in range(40, 50):  # Virga Jesse ---> geen litterals!

                # maak dict wachten met key(functie) uit col 2 en value (staflid) uit col en row
                wachten[sheet.cell(row, 2).value] = sheet.cell(row, col).value

            # for row in range(120, 130):  # Salvator
            #
            #     # maak dict wachten met key(functie) uit col 2 en value (staflid) uit col en row
            #     wachten[sheet.cell(row, 2).value] = sheet.cell(row, col).value
            # maak dictionary voor S1 Salvator
            for row in range(100, 150):
                if sheet.cell(row, 2).value == 'S1':
                    wachten['S1'] = sheet.cell(row, col).value

            if not (wachten['V1 (Coronapermanentie)'] in intensieve or wachten['V2 (INSLAPEND)'] in intensieve or
                    wachten['S1'] in intensieve):
                result.append(f'Check Wachten op {date}')
    if len(result) == 0:
        result.append('Geen problemen gevonden in de wachtsamenstelling')
    return result


def checkrecup(schemas):
    # Indien recup ingepland dan mag schema alleen trigram, dag, datum en R1 of RECUP bevatten
    result = list()
    for schema in schemas:
        if ("R1" in schema or "RECUP" in schema) and not checkweekend(schema[1]):
            if len(schema) > 4:
                result.append(f'Controleer recuperatie voor: {schema[0]} op {schema[2]}')
    if len(result) == 0:
        result.append('Recuperaties in orde')
    return result


def check_v3(schemas):
    result = list()
    for i in range(0, len(schemas)):
        if "V3 (THUISWACHT)" in schemas[i]:
            if not (schemas[i][1] == 'Vrijdag' or schemas[i][1] == 'Zaterdag'):
                if not ("V16" in schemas[i + 1] or "SNIPPERDAG" in schemas[i + 1] or "Onverwachte Snipperdag" in
                        schemas[i + 1]):
                    result.append(f'Controleer positie na V3 van {schemas[i][0]} op {schemas[i][2]}')
    if len(result) == 0:
        result.append('Posities na V3 in orde.')
    return result


def check_aanwezigen_sal_opdatum(weken):
    aanwezigen = {}
    for sheet in weken:
        for col in range(3, 8):

            temp_date = sheet.cell(4, col).value.date()
            date = temp_date.strftime("%d/%m/%Y")
            aanwezigen[date] = []

            for row in range(MIN_SAL, MAX_SAL + 1):
                aanwezigen[date].append(sheet.cell(row, col).value)

    return aanwezigen


def check_locoreg_anesthesist_sal(aanwezigen, lijst):
    """loop over dict aanwezigen en zoek naar locoregionale anesthesist in values, minstens één nodig"""
    result = list()
    for i in aanwezigen:
        counter = 0
        for el in aanwezigen[i]:
            if el in lijst["locoreg_anesthesist"]:
                counter += 1
        if not counter > 0:
            result.append(f'Geen locoregionale anesthesist in Salvator op {i}')
    return result


def check_aanwezigen_virga_var_opdatum(weken, aantal):
    """Maak dict met key = datum en eerste aantal anesthesisten in Virga"""
    aanwezigen = {}
    for sheet in weken:

        # Bepaal het startpunt in de kolom met posities, V2 is stabielste parameter
        for row in range(1, sheet.max_row):
            if sheet.cell(row, 2).value == "V2 (INSLAPEND)":
                startpunt = row - 1

        # maak key aan en initieer value list
        for col in range(3, 8):
            temp_date = sheet.cell(4, col).value.date()
            date = temp_date.strftime("%d/%m/%Y")
            aanwezigen[date] = []

            # vul de value list
            for row in range(startpunt, startpunt + aantal):
                aanwezigen[date].append(sheet.cell(row, col).value)

    return aanwezigen


def check_cardioanesthesist_vj(aanwezigen, lijst):
    result = list()
    for i in aanwezigen:
        counter = 0
        for el in aanwezigen[i]:
            if el in lijst["cardio_anesthesist"]:
                counter += 1
        if not counter >= 2:
            result.append(f'Te weinig cardio anesthesisten in positie V1 tot V7 op {i}')
    return result











